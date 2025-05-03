import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox, filedialog
import mysql.connector
import pandas as pd
from PIL import Image
import subprocess
import os
from datetime import datetime

# App config
ctk.set_appearance_mode("light")
admin_dashboard = ctk.CTk()
admin_dashboard.geometry("1111x750")
admin_dashboard.title("Admin Dashboard")
admin_dashboard.configure(fg_color="#7F5B6A")

# Top Frame
top_frame = ctk.CTkFrame(admin_dashboard, height=60, fg_color="#D9D9D9", corner_radius=0)
top_frame.pack(fill="x")
title_label = ctk.CTkLabel(top_frame, text="ADMIN DASHBOARD", font=("Arial", 22, "bold"), text_color="black")
title_label.place(relx=0.03, rely=0.5, anchor="w")

# Image utility
def resize_image(size, image_url):
    image = Image.open(image_url)
    return ctk.CTkImage(light_image=image, size=size)

# Navigation
def go_home():
    subprocess.Popen(["python", "HomePage.py"])
    admin_dashboard.destroy()

def open_current_event():
    subprocess.Popen(["python", "CurrentEvent.py"])
    admin_dashboard.destroy()

def open_staff_allocation():
    subprocess.Popen(["python", "StaffAllocation.py"])
    admin_dashboard.destroy()

# Back and Logout Icons
logout_icon = resize_image((35, 35), "icons/door.png")
logout_label = ctk.CTkLabel(top_frame, text="", image=logout_icon, cursor="hand2")
logout_label.place(x=1060, y=10)
logout_label.bind("<Button-1>", lambda e: go_home())

# Dashboard Grid
frame1 = ctk.CTkFrame(admin_dashboard, width=284, height=244, fg_color="#D9D9D9", corner_radius=10)
frame1.place(x=80, y=120)
cevent_icon = resize_image((95, 95), "icons/Current Events.png")
cevent_btn = ctk.CTkButton(frame1, text="", image=cevent_icon, fg_color="#D9D9D9", width=95, height=95, command=open_current_event)
cevent_btn.place(x=95, y=30)
cevent_label = ctk.CTkLabel(frame1, text="Current Events", text_color="#000000", font=('inter', 24))
cevent_label.place(x=70, y=180)

frame2 = ctk.CTkFrame(admin_dashboard, width=284, height=244, fg_color="#D9D9D9", corner_radius=10)
frame2.place(x=400, y=120)
staff_icon = resize_image((95, 95), "icons/Staff Allocation.png")
staff_btn = ctk.CTkButton(frame2, text="", image=staff_icon, fg_color="#D9D9D9", width=95, height=95, command=open_staff_allocation)
staff_btn.place(x=95, y=30)
staff_label = ctk.CTkLabel(frame2, text="Staff Allocation", text_color="#000000", font=('inter', 24))
staff_label.place(x=55, y=180)

# Fetch events
def fetch_events():
    try:
        conn = mysql.connector.connect(
            host="141.209.241.57",
            user="tiruv1h",
            password="mypass",
            database="BIS698W1830_GRP1"
        )
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM events")
        events = cursor.fetchall()
        conn.close()
        return events
    except Exception as e:
        messagebox.showerror("Database Error", str(e))
        return []

# Generate Report Function with timestamped filename and cleaned format
from openpyxl.styles import Font, Alignment, PatternFill

def generate_report():
    events = fetch_events()
    if not events:
        messagebox.showwarning("No Events", "No event data available.")
        return

    df = pd.DataFrame(events)

    # Format date and drop ID column
    df['date'] = pd.to_datetime(df['date'], errors='coerce')
    df['date'] = df['date'].dt.strftime('%d %B %Y')
    if 'id' in df.columns:
        df.drop(columns=['id'], inplace=True)

    upcoming = df[df['date'] >= datetime.now().strftime('%d %B %Y')].shape[0]
    total_events = df.shape[0]
    statuses = df['status'].value_counts().to_dict() if 'status' in df.columns else {}

    summary_data = {
        "Metric": ["Total Events", "Upcoming Events"] + list(statuses.keys()),
        "Value": [total_events, upcoming] + list(statuses.values())
    }
    summary_df = pd.DataFrame(summary_data)

    # Timestamped filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    default_filename = f"Admin_Event_Report_{timestamp}.xlsx"

    base_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        initialfile=default_filename,
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not base_path:
        return

    with pd.ExcelWriter(base_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary Report", index=False, startrow=2)
        df.to_excel(writer, sheet_name="All Events Data", index=False, startrow=2)

        workbook = writer.book
        summary_sheet = workbook["Summary Report"]
        events_sheet = workbook["All Events Data"]

        # Define style
        heading_font = Font(bold=True, size=14, color="FFFFFF")
        heading_alignment = Alignment(horizontal="center", vertical="center")
        heading_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # Apply to Summary Report
        summary_sheet["A1"] = "Admin Event Summary Report"
        summary_sheet.merge_cells("A1:B1")
        summary_sheet["A1"].font = heading_font
        summary_sheet["A1"].alignment = heading_alignment
        summary_sheet["A1"].fill = heading_fill

        # Apply to All Events Data
        events_sheet["A1"] = "Detailed Event Data"
        events_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        events_sheet["A1"].font = heading_font
        events_sheet["A1"].alignment = heading_alignment
        events_sheet["A1"].fill = heading_fill

    messagebox.showinfo("Success", f"Report saved to:\n{base_path}")

# Generate Button
generate_button = ctk.CTkButton(
    admin_dashboard, text="📥 Generate Reports", font=("Arial", 13, "bold"),
    command=generate_report, width=160, height=35
)
generate_button.place(x=920, y=85)

admin_dashboard.mainloop()
